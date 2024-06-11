from openpyxl import Workbook
import re

# 打开文件并读取内容
with open('D:/physionet/222.txt', 'r') as file:
    content = file.read()
 
# 提取数据
extracted_data = content.split('\n')  # 假设数据是按行分隔的

# 处理提取的数据
i=1
a=[]
b=[]
for data in extracted_data:# 1 2 4 5 7 8 
    #if((i-1)%3==0):
    #    a.append(data)
    if((i-2)%3==0):
        #print(float(re.findall(r'\d+', data)[1])/1e4)
        b.append(float(re.findall(r'\d+', data)[1])/1e4)
    #print(data)
    i=i+1
print(b)
# 创建一个新的工作簿
wb = Workbook()
 
# 选择活动的工作表
ws = wb.active
 
# 要写入的数据列表
 
# 写入数据到第一列（A列）
j=1
for row in range(1, len(b) + 1):
    ws.cell(row=row, column=1, value=j)
    ws.cell(row=row, column=2, value=b[row - 1])
    j=j+1
# 保存工作簿
wb.save('D:/physionet/222.xlsx')
